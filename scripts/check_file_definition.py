#!/usr/bin/env python3
"""
⑤チェック: 入力CSVのカラム構成がファイル定義書(fileteigisyo_YYYY.xlsx)と一致するか

使い方:
  python scripts/check_file_definition.py --year 2024
  python scripts/check_file_definition.py --year 2022 2023 2024

2019-2021はxlsx未存在（PDFのみ）のためスキップ。

終了コード: 不一致あり=1, 一致=0
"""
import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).parent.parent
KNOWN_YEARS_WITH_XLSX = list(range(2022, 2025))


def get_xlsx_columns(xlsx_path, sheet='本票'):
    """ファイル定義書xlsxから本票の項目名リストを抽出"""
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    cols = []
    for row in ws.iter_rows(values_only=True):
        no = row[0]
        name = row[1]
        if no is None or not str(no).strip().isdigit():
            continue
        if name is None or str(name).strip() == '':
            continue
        cols.append(str(name).strip())
    return cols


def get_csv_columns(csv_path):
    """実CSV（Shift-JIS）のヘッダーカラム名リストを返す"""
    df = pd.read_csv(csv_path, dtype=object, encoding='cp932', nrows=0)
    return list(df.columns)


def check_year(year):
    xlsx_path = ROOT / 'data' / str(year) / f'fileteigisyo_{year}.xlsx'
    csv_path = ROOT / 'data' / str(year) / f'honhyo_{year}.csv'

    if not xlsx_path.exists():
        print(f'  [SKIP] fileteigisyo_{year}.xlsx が存在しません（2019-2021はPDFのみ）')
        return True

    if not csv_path.exists():
        print(f'  [SKIP] honhyo_{year}.csv が存在しません')
        return True

    xlsx_cols = get_xlsx_columns(xlsx_path, '本票')
    if xlsx_cols is None:
        print(f'  [SKIP] fileteigisyo_{year}.xlsx に「本票」シートがありません')
        return True

    csv_cols = get_csv_columns(csv_path)

    if xlsx_cols == csv_cols:
        print(f'  ✅ {year}: カラム構成一致（{len(xlsx_cols)}列）')
        return True

    # 差異あり
    print(f'  ❌ {year}: カラム構成不一致')
    print(f'     ファイル定義書: {len(xlsx_cols)}列 / 実CSV: {len(csv_cols)}列')

    only_xlsx = [c for c in xlsx_cols if c not in csv_cols]
    only_csv = [c for c in csv_cols if c not in xlsx_cols]
    order_diff = [
        (i, xlsx_cols[i], csv_cols[i])
        for i in range(min(len(xlsx_cols), len(csv_cols)))
        if xlsx_cols[i] != csv_cols[i]
    ]

    for c in only_xlsx[:10]:
        print(f'     xlsx のみ: {c!r}')
    for c in only_csv[:10]:
        print(f'     csv  のみ: {c!r}')
    if order_diff:
        print(f'     位置ずれ（列番号 xlsx → csv）:')
        for idx, xc, cc in order_diff[:5]:
            print(f'       col {idx+1}: xlsx={xc!r} / csv={cc!r}')

    return False


def main():
    parser = argparse.ArgumentParser(description='⑤チェック: 入力CSVとファイル定義書の列構成確認')
    parser.add_argument('--year', type=int, nargs='+', help='対象年（例: 2022 2023 2024）')
    parser.add_argument('--all', action='store_true', help='全年次チェック')
    args = parser.parse_args()

    if args.all:
        years = list(range(2019, 2025))
    else:
        years = args.year or [2024]

    ok = True
    for year in years:
        print(f'\n=== {year}年 ===')
        if not check_year(year):
            ok = False

    print(f'\n{"="*60}')
    if ok:
        print('✅ 全年次: カラム構成一致（またはスキップ）')
    else:
        print('❌ カラム構成の不一致があります。decoder/convert を修正してください。')
        sys.exit(1)


if __name__ == '__main__':
    main()
