#!/usr/bin/env python3
"""
①チェック: 公式コード表(xlsx) vs 変換に使用するコード表（CSV + Pythonコード辞書）

使い方:
  python scripts/check_codetable_vs_official.py --year 2024
  python scripts/check_codetable_vs_official.py --year 2022 2023 2024
  python scripts/check_codetable_vs_official.py --all

終了コード: 差異あり=1, 全て一致=0
"""
import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
CODE_PATTERN = re.compile(r'^[0-9A-Za-z]{1,6}$')


# ---- xlsx抽出ユーティリティ ----

def _is_code_value(val):
    """短いコード値かどうか判定（ヘッダー行や説明文を除外）"""
    if val is None:
        return False
    return bool(CODE_PATTERN.match(str(val).strip()))


def extract_xlsx_codes(ws, key_col_indices, label_col):
    """
    xlsxシートからコード辞書を抽出。
    key_col_indices: xlsxの列インデックス (0始まり)
    データ行はkey_col_indices[0]の値が短いコード文字列の行のみ収集。
    単一キー列でカンマ区切り ('12,14') の場合は個別エントリに展開する。
    """
    codes = {}
    for row in ws.iter_rows(values_only=True):
        first_key_val = row[key_col_indices[0]] if len(row) > key_col_indices[0] else None
        # カンマ区切り複合コード ('12,14') の先頭部分もコードとして認識
        first_str = str(first_key_val).strip().split(',')[0].strip() if first_key_val is not None else ''
        if not _is_code_value(first_str):
            continue
        try:
            raw_label = row[label_col] if row[label_col] is not None and row[label_col] != '' else ''
            label = str(raw_label).replace('\n', '').strip()
            if len(key_col_indices) == 1:
                # 単一キー列: カンマ区切りを展開
                for part in str(first_key_val).split(','):
                    part = part.strip()
                    if part:
                        codes[part] = label
            else:
                key = '|'.join(
                    str(row[c]).strip() if row[c] is not None else ''
                    for c in key_col_indices
                )
                if key.replace('|', '').strip():
                    codes[key] = label
        except IndexError:
            continue
    return codes


def extract_csv_codes(filepath, key_col_indices, label_col, skip_rows=5):
    """CSVからコード辞書を抽出"""
    codes = {}
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for _ in range(skip_rows):
            next(reader, None)
        for row in reader:
            if not row or all(not c.strip() for c in row):
                continue
            first_key_val = row[key_col_indices[0]] if len(row) > key_col_indices[0] else None
            if not _is_code_value(first_key_val):
                continue
            try:
                key = '|'.join(str(row[c]).strip() for c in key_col_indices)
                label = str(row[label_col]).strip() if len(row) > label_col else ''
                if key.replace('|', '').strip():
                    codes[key] = label
            except IndexError:
                continue
    return codes


def dict_to_codes(d):
    """Pythonコード辞書を {str(key): label} 形式に正規化"""
    return {str(k).strip(): str(v).strip() for k, v in d.items()}


def compare_dicts(xlsx_codes, target_codes, name):
    """2つのコード辞書を比較し差異リストを返す"""
    errors = []
    only_xlsx = {k: v for k, v in xlsx_codes.items() if k not in target_codes}
    only_target = {k: v for k, v in target_codes.items() if k not in xlsx_codes}
    label_diff = {
        k: (xlsx_codes[k], target_codes[k])
        for k in xlsx_codes
        if k in target_codes and xlsx_codes[k] != target_codes[k]
    }
    if only_xlsx or only_target or label_diff:
        errors.append(f'  [DIFF] {name}')
        for k, v in sorted(only_xlsx.items())[:10]:
            errors.append(f'    xlsx のみ: {k!r:>12} = {v!r}')
        for k, v in sorted(only_target.items())[:10]:
            errors.append(f'    変換用 のみ: {k!r:>12} = {v!r}')
        for k, (xv, cv) in sorted(label_diff.items())[:10]:
            errors.append(f'    ラベル差異: {k!r:>12}: xlsx={xv!r} / 変換用={cv!r}')
        if len(only_xlsx) > 10:
            errors.append(f'    ... 他 {len(only_xlsx)-10} 件')
    else:
        print(f'  ✅ {name:40s} ({len(xlsx_codes)}件 一致)')
    return errors


# ---- Part A: CSVファイル vs xlsx ----
# コンバーターが _load_csv_dict で直接読み込む3種のCSV

CSV_CHECKS = [
    # (sheet_name, xlsx_key_cols, xlsx_label_col, csv_path_tmpl, csv_key_cols, csv_label_col)
    ('都道府県',      [1], 2, 'common/2_koudohyou_todouhukenkoudo.csv',                 [0], 1),
    ('警察署等',      [1, 2], 4, '{year}/3_koudohyou_keisatusyotoukoudo.csv',            [0, 1], 3),
    ('路線 (高速)',   [1, 2], 4, '{year}/9_koudohyou_rosen_kousokujidousyasenyou.csv',   [0, 1], 3),
    ('トンネル番号',  [1, 2, 3], 4, '{year}/53_koudohyou_tonnerubangou.csv',             [0, 1, 2], 3),
]


def check_csv_vs_xlsx(wb, year, errors):
    for sheet_name, xkey, xlabel, csv_tmpl, ckey, clabel in CSV_CHECKS:
        if sheet_name not in wb.sheetnames:
            continue
        csv_rel = csv_tmpl.replace('{year}', str(year))
        csv_path = ROOT / 'code_tables' / csv_rel
        if not csv_path.exists():
            errors.append(f'  [MISSING CSV] {csv_path.relative_to(ROOT)}')
            continue
        xlsx_codes = extract_xlsx_codes(wb[sheet_name], xkey, xlabel)
        csv_codes = extract_csv_codes(csv_path, ckey, clabel)
        errs = compare_dicts(xlsx_codes, csv_codes, f'{sheet_name} (CSV: {csv_path.relative_to(ROOT)})')
        errors.extend(errs)


# ---- Part B: Pythonコード辞書 vs xlsx ----
# converter/codes/yYYYY.py の各辞書名 → xlsxシート名 の対応

DICT_TO_SHEET = [
    # (dict_attr, sheet_name, xlsx_key_cols, xlsx_label_col)
    ('TYUUYA',                  '昼夜',                 [1], 2),
    ('TENKOU',                  '天候',                 [1], 2),
    ('TIKEI',                   '地形',                 [1], 2),
    ('ROMENJYOUTAI',             '路面状態',              [1], 2),
    ('DOUROKEIJYOU',             '道路形状',              [1], 2),
    ('SINGOUKI',                '信号機',                [1], 2),
    ('ITIJITEISIKISEI_HYOUSIKI', '一時停止規制 標識',      [1], 2),
    ('ITIJITEISIKISEI_HYOUJI',   '一時停止規制 表示',      [1], 2),
    ('SYADOUHUKUIN',             '車道幅員',              [1], 2),
    ('DOUROSENKEI',              '道路線形',              [1], 2),
    ('SYOUTOTUTITEN',            '衝突地点',              [1], 2),
    ('ZOONKISEI',               'ゾーン規制',             [1], 2),
    ('TYUUOUBUNRITAISISETUTOU',  '中央分離帯施設等',       [1], 2),
    ('HOSYADOUKUBUN',            '歩車道区分',             [1], 2),
    ('JIKORUIKEI',              '事故類型（本票）',        [1], 2),
    ('NENREI',                  '年齢',                  [1], 2),
    ('TOUJISYASYUBETU',         '当事者種別',             [1], 2),
    ('YOUTOBETU',               '用途',                  [1], 2),
    ('SYARYOUKEIJYOU',          '車両形状',               [1], 2),
    ('SOKUDOKISEI_SITEINOMI',   '速度規制（指定のみ）',    [1], 2),
    ('SYARYOUNOSONKAITEIDO',    '車両の損壊程度',          [1], 2),
    ('EABAGUNOSOUBI',           'エアバッグの装備',         [1], 2),
    ('SAIDOEABAGUNOSOUBI',      'サイドエアバッグの装備',   [1], 2),
    ('JINSINSONSYOUTEIDO',      '人身損傷程度',            [1], 2),
    ('YOUBI',                   '曜日',                  [1], 2),
    ('SYUKUJITU',               '祝日',                  [1], 2),
]

# xlsxにあるが変換用辞書に不要なキー（複合コードの片方など）
XLSX_ONLY_ALLOWED = {
    # 車道幅員: xlsx は '12,14'→1ラベルのように複合コードを1行で表現。
    # 変換データでは '12','13','16' は実際には出現せず、'14','17','18' のみ使用。
    'SYADOUHUKUIN': {'12', '13', '16'},
}

# xlsx側で「00: 対象外当事者」などが存在しない場合がある（変換用独自エントリ）
DICT_ONLY_ALLOWED = {
    # dict_attr → 変換用辞書にのみ存在が許容されるキーのセット（xlsxになくてもOK）
    'TOUJISYASYUBETU':          {'00', '71', '72', '75', '76'},
    'YOUTOBETU':                {'00'},
    'SYARYOUKEIJYOU':           {'00'},
    'SOKUDOKISEI_SITEINOMI':    {'00'},
    'SYARYOUNOSONKAITEIDO':     {'0', ''},
    'EABAGUNOSOUBI':            {'0'},
    'SAIDOEABAGUNOSOUBI':       {'0'},
    'JINSINSONSYOUTEIDO':       {'0'},
    'ITIJITEISIKISEI_HYOUSIKI': {'00'},
    'ITIJITEISIKISEI_HYOUJI':   {'00'},
    'DOUROSENKEI':              {'0'},
    'TYUUOUBUNRITAISISETUTOU':  {'0'},
    'HOSYADOUKUBUN':            set(),
    'DOUROKEIJYOU':             {'00'},
    'SINGOUKI':                 set(),
    'NENREI':                   {'00'},
}


def check_dict_vs_xlsx(wb, year, errors):
    from converter.codes import get_codes
    c = get_codes(year)

    for dict_attr, sheet_name, xkey, xlabel in DICT_TO_SHEET:
        if sheet_name not in wb.sheetnames:
            continue
        if not hasattr(c, dict_attr):
            errors.append(f'  [MISSING DICT] {dict_attr} が codes モジュールにありません')
            continue

        xlsx_codes = extract_xlsx_codes(wb[sheet_name], xkey, xlabel)
        py_codes = dict_to_codes(getattr(c, dict_attr))
        allowed_only = DICT_ONLY_ALLOWED.get(dict_attr, set())

        xlsx_allowed = XLSX_ONLY_ALLOWED.get(dict_attr, set())
        only_xlsx = {k: v for k, v in xlsx_codes.items() if k not in py_codes and k not in xlsx_allowed}
        only_py = {k: v for k, v in py_codes.items() if k not in xlsx_codes and k not in allowed_only}
        label_diff = {
            k: (xlsx_codes[k], py_codes[k])
            for k in xlsx_codes
            if k in py_codes and xlsx_codes[k] != py_codes[k]
        }

        name = f'{dict_attr} vs xlsx[{sheet_name}]'
        if only_xlsx or only_py or label_diff:
            errors.append(f'  [DIFF] {name}')
            for k, v in sorted(only_xlsx.items())[:10]:
                errors.append(f'    xlsx のみ: {k!r:>12} = {v!r}')
            for k, v in sorted(only_py.items())[:10]:
                errors.append(f'    辞書 のみ: {k!r:>12} = {v!r}')
            for k, (xv, cv) in sorted(label_diff.items())[:10]:
                errors.append(f'    ラベル差異: {k!r:>12}: xlsx={xv!r} / 辞書={cv!r}')
        else:
            print(f'  ✅ {name:50s} ({len(xlsx_codes)}件 一致)')


# ---- メイン ----

def check_year(year):
    xlsx_path = ROOT / 'data' / str(year) / f'codebook_{year}.xlsx'
    if not xlsx_path.exists():
        print(f'  [SKIP] codebook_{year}.xlsx が存在しません')
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    errors = []

    print(f'  --- Part A: CSV vs xlsx ---')
    check_csv_vs_xlsx(wb, year, errors)

    print(f'  --- Part B: Pythonコード辞書 vs xlsx ---')
    check_dict_vs_xlsx(wb, year, errors)

    return errors


def main():
    parser = argparse.ArgumentParser(description='①チェック: 公式コード表(xlsx) vs 変換用コード表')
    parser.add_argument('--year', type=int, nargs='+', help='対象年（例: 2022 2023 2024）')
    parser.add_argument('--all', action='store_true', help='2022〜2024すべてチェック')
    args = parser.parse_args()

    years = list(range(2022, 2025)) if args.all else (args.year or [2024])

    all_errors = []
    for year in years:
        print(f'\n=== {year}年 ===')
        errs = check_year(year)
        if errs:
            all_errors.extend([f'[{year}] ' + e.strip() for e in errs])
            for e in errs:
                print(e)
        else:
            print(f'  ✅ {year}: 全コード表一致')

    print(f'\n{"="*60}')
    if all_errors:
        print(f'❌ {len(all_errors)} 件の不一致があります')
        sys.exit(1)
    else:
        print('✅ 全年次: 公式xlsx と 変換用コード表 は完全一致')
        sys.exit(0)


if __name__ == '__main__':
    main()
