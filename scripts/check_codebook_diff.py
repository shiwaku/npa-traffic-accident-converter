#!/usr/bin/env python3
"""
年次間のコード表差分チェックスクリプト

使い方:
  python scripts/check_codebook_diff.py --base 2024 --new 2025
"""
import argparse
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent


def extract_codes(ws):
    """xlsxシートからコード→ラベルの辞書を抽出（先頭4行はヘッダーとしてスキップ）"""
    codes = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        cells = [c for c in row if c is not None]
        if len(cells) >= 2:
            code = str(cells[0]).strip()
            label = str(cells[1]).strip()
            if code:
                codes[code] = label
    return codes


def compare_codebooks(base_year: int, new_year: int):
    base_path = ROOT / 'data' / str(base_year) / f'codebook_{base_year}.xlsx'
    new_path = ROOT / 'data' / str(new_year) / f'codebook_{new_year}.xlsx'

    if not base_path.exists():
        print(f"ERROR: {base_path} が存在しません")
        return
    if not new_path.exists():
        print(f"ERROR: {new_path} が存在しません")
        print(f"  → data/{new_year}/codebook_{new_year}.xlsx をダウンロードしてから実行してください")
        return

    base_wb = openpyxl.load_workbook(base_path)
    new_wb = openpyxl.load_workbook(new_path)

    print(f"\n{'='*60}")
    print(f" コード表差分チェック: {base_year} → {new_year}")
    print(f"{'='*60}")

    all_sheets = sorted(set(base_wb.sheetnames) | set(new_wb.sheetnames))
    has_diff = False

    for sheet_name in all_sheets:
        if sheet_name not in base_wb.sheetnames:
            print(f"\n[NEW SHEET] {sheet_name}")
            has_diff = True
            continue
        if sheet_name not in new_wb.sheetnames:
            print(f"\n[DELETED SHEET] {sheet_name}")
            has_diff = True
            continue

        base_codes = extract_codes(base_wb[sheet_name])
        new_codes = extract_codes(new_wb[sheet_name])

        added = {k: v for k, v in new_codes.items() if k not in base_codes}
        removed = {k: v for k, v in base_codes.items() if k not in new_codes}
        changed = {
            k: (base_codes[k], new_codes[k])
            for k in new_codes
            if k in base_codes and base_codes[k] != new_codes[k]
        }

        if added or removed or changed:
            has_diff = True
            print(f"\n[変更あり] {sheet_name}  ({len(base_codes)}件 → {len(new_codes)}件)")
            for k, v in sorted(added.items()):
                print(f"  + 追加  {k!s:>6}: {v}")
            for k, v in sorted(removed.items()):
                print(f"  - 削除  {k!s:>6}: {v}")
            for k, (old, new) in sorted(changed.items()):
                print(f"  ~ 変更  {k!s:>6}: {old!r} → {new!r}")

    if not has_diff:
        print(f"\n✅ 差分なし（{base_year}年と{new_year}年のコード表は同一）")
    else:
        print(f"\n⚠️  上記の差分を converter/codes/y{new_year}.py に反映してください")
        print(f"   参考: converter/codes/y{base_year}.py をコピーして差分のみ修正")


def main():
    parser = argparse.ArgumentParser(description='年次間のコード表差分チェック')
    parser.add_argument('--base', type=int, required=True, help='比較元の年（例: 2024）')
    parser.add_argument('--new', type=int, required=True, help='新しい年（例: 2025）')
    args = parser.parse_args()
    compare_codebooks(args.base, args.new)


if __name__ == '__main__':
    main()
